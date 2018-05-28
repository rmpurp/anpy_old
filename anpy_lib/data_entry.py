#!/usr/bin/env python3

import datetime
from openpyxl import Workbook
from anpy_lib import column_creation


class Datum:
    def __init__(self, date, value, column_class):
        self.date = date
        self.value = value
        self.column_class = column_class

    def add(self, ws, columns):
        col = self.get_col(columns)
        row = self.get_row(ws, columns)
        ws.cell(row=row, column=col, value=self.value)

    def get_row(self, ws, columns):
        start_row = 2
        date_column = get_col_from_class(columns, column_creation.DateColumn)
        base_date = ws.cell(row=start_row, column=date_column).value
        date_difference = (self.date - base_date).days
        return start_row + date_difference

    def get_col(self, columns):
        return get_col_from_class(columns, self.column_class)


def get_col_from_class(columns, col_class):
    return columns.index(col_class.__name__) + 1


def get_col(column_name, columns):
    return columns.index(column_name) + 1


def get_row(ws, date, columns):
    start_row = 2
    date_column = get_col_from_class(columns, column_creation.DateColumn)
    base_date = ws.cell(row=start_row, column=date_column).value
    if isinstance(base_date, datetime.datetime):
        base_date = base_date.date()
    date_difference = (date - base_date).days
    return start_row + date_difference


def add_data(ws, date, value, column_name, columns, cell_format=None):
    row = get_row(ws, date, columns)
    col = get_col(column_name, columns)
    cell = ws.cell(row=row, column=col, value=value)
    if cell_format:
        cell.number_format = cell_format


def add_list(ws, date, values, columns):
    row = get_row(ws, date, columns)
    start_col = len(columns) + 1
    for idx, value in enumerate(values):
        ws.cell(row=row, column=start_col + idx, value=value)


if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active

    subjects = ['AP Euro', 'AP Bio', 'Band', 'Dab', 'dab2']

    column_creation.create_stat_columns()
    column_creation.create_subjects(subjects)
    column_creation.Column.make_all(ws)
    columns = column_creation.Column.get_column_strings()
    time_started = Datum(datetime.date(2018, 5, 22), datetime.time(13, 1),
                         column_creation.TimeStartedColumn)
    time_ended = Datum(datetime.date(2018, 5, 22), datetime.time(15, 1),
                       column_creation.TimeEndedColumn)

    time_started.add(ws, columns)
    time_ended.add(ws, columns)

    wb.save('test.xlsx')
