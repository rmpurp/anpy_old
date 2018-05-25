from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import itertools as it
from openpyxl.worksheet.cell_range import CellRange
from anpy_lib.utils import get_most_recent_monday


NUM_BODY_ITEMS = 7
"""Number of rows in a column between the header row and the footer row"""


class Column:
    col_order = []

    def __init__(self, title: str, dependencies=None):
        '''A column in an Excel spreadsheet.

        A column has a title row, several body rows, and a footer row.
        '''
        self.title = title
        Column.col_order.append(self)
        self.column = len(Column.col_order)
        if not dependencies:
            dependencies = []
        self._dep = dict(zip(dependencies, it.repeat(-1)))

    def _satisfy_dependencies(self):
        '''Find the columns needed upon and store internally.'''
        for col_type in self._dep:
            self._dep[col_type] = Column.find_column_of_type(col_type)

    def make(self, ws):
        '''Makes this column in the given worksheet'''
        self._satisfy_dependencies()
        ws.cell(row=1, column=self.column, value=self.title)
        for row, item in enumerate(self._get_body(NUM_BODY_ITEMS), 2):
            cell = ws.cell(row=row, column=self.column, value=item)
            self._body_cell_op(cell)
        ws.cell(row=row + 1, column=self.column, value=self._get_footer())

    def _get_body_item(self, item_num):
        '''Get the item_num-th body item'''
        return item_num

    def _body_cell_op(self, cell):
        '''Callback function that is called when body cell is made.

        To be overriden by subclasses.

        Args:
            cell: the body cell that was just created
        '''
        pass

    def _get_body(self, num_items):
        '''Yield the body cells.

        Args:
            num_items: the number of body items
        '''
        i = 1
        while i <= num_items:
            yield self._get_body_item(i)
            i += 1

    def _get_footer(self):
        range = CellRange(min_col=self.column,
                          max_col=self.column,
                          min_row=2,
                          max_row=1 + NUM_BODY_ITEMS)
        return '=AVERAGE({})'.format(str(range))

    def __str__(self):
        return type(self).__name__

    @staticmethod
    def find_column_of_type(col_type):
        return [isinstance(c, col_type)
                for c in Column.col_order].index(True) + 1

    @staticmethod
    def make_all(worksheet):
        for c in Column.col_order:
            c.make(worksheet)

    @staticmethod
    def get_column_strings():
        return [str(c) for c in Column.col_order]


class DateColumn(Column):
    def __init__(self, start_date=None):
        if not start_date:
            start_date = get_most_recent_monday()
        self.start_date = start_date
        super().__init__('Date')

    def _get_body_item(self, item_num):
        if item_num == 1:
            return self.start_date
        else:
            return '={}{} + 1'.format(get_column_letter(self.column), item_num)

    def _body_cell_op(self, cell):
        cell.number_format = 'YYYY-MM-DD'

    def _get_footer(self):
        return 'Averages:'


class DefaultValueColumn(Column):
    def __init__(self, title, default_value='N/A'):
        self.default_value = default_value
        super().__init__(title)

    def _get_body_item(self, item_num):
        return self.default_value


class DataColumn(DefaultValueColumn):
    def __init__(self, title):
        super().__init__(title, default_value=0)

    def _body_cell_op(self, cell):
        cell.number_format = '0'


class TimeStartedColumn(DefaultValueColumn):
    def __init__(self):
        super().__init__('Time Started')

    def _get_footer(self):
        return 'N/A'


class TimeEndedColumn(DefaultValueColumn):
    def __init__(self):
        super().__init__('Time Ended')

    def _get_footer(self):
        return 'N/A'


class TimeTotalColumn(Column):
    def __init__(self):
        dep = [TimeStartedColumn, TimeEndedColumn]
        # self.time_started_col = get_column_letter(time_started_col)
        # self.time_ended_col = get_column_letter(time_ended_col)
        super().__init__('Time Total (H)', dependencies=dep)

    def _get_body_item(self, item_num):
        row = item_num + 1
        time_started_col = get_column_letter(self._dep[TimeStartedColumn])
        time_ended_col = get_column_letter(self._dep[TimeEndedColumn])
        template = '=IF({0}{1}="N/A","N/A",' \
            + ' (MOD(24 + (60 * HOUR({2}{1}) - 60 * HOUR({0}{1})' \
            + ' + MINUTE({2}{1}) - MINUTE({0}{1})) / 60, 24)))'
        return template.format(time_started_col, row, time_ended_col)

    def _body_cell_op(self, cell):
        cell.number_format = '0.0'


class TimeWorkingColumn(Column):
    def __init__(self):
        dep = [DataColumn]
        super().__init__('Time Working (H)', dependencies=dep)

    def _get_body_item(self, item_num):
        data_start_idx = self._dep[DataColumn]
        data_end_idx = len(Column.col_order)
        row = item_num + 1
        template = '=IF(SUM({0})=0,"N/A",SUM({0})/60)'
        cell_range = CellRange(min_col=data_start_idx,
                               max_col=data_end_idx,
                               min_row=row,
                               max_row=row)
        return template.format(cell_range)

    def _body_cell_op(self, cell):
        cell.number_format = '0.0'


class EfficiencyColumn(Column):
    def __init__(self):
        dep = [TimeTotalColumn, TimeWorkingColumn]
        super().__init__('% Efficiency', dependencies=dep)

    def _get_body_item(self, item_num):
        row = item_num + 1
        time_total_col = self._dep[TimeTotalColumn]
        time_working_col = self._dep[TimeWorkingColumn]
        template = '=IF({2}{1}="N/A","N/A",IFERROR({2}{1}/({0}{1}*0.75),0))'
        return template.format(get_column_letter(time_total_col),
                               row,
                               get_column_letter(time_working_col))

    def _get_footer(self):
        time_total_col = self._dep[TimeTotalColumn]

        total_range = CellRange(min_col=time_total_col,
                                max_col=time_total_col,
                                min_row=2,
                                max_row=NUM_BODY_ITEMS + 1)
        my_range = CellRange(min_col=self.column,
                             max_col=self.column,
                             min_row=2,
                             max_row=NUM_BODY_ITEMS + 1)
        template = '=SUMPRODUCT({0},{1}) / SUM({0})'
        return template.format(total_range, my_range)

    def _body_cell_op(self, cell):
        cell.number_format = '0.0%'


DEFAULT_COLUMNS = [DateColumn,
                   TimeStartedColumn,
                   TimeEndedColumn,
                   TimeTotalColumn,
                   TimeWorkingColumn,
                   EfficiencyColumn]


def get_subjects(ws, num_titles):
    subjects = []
    index = num_titles + 1
    cell = ws.cell(row=1, column=index)
    while cell.value:
        subjects.append(cell.value)
        index += 1
        cell = ws.cell(row=1, column=index)
    return subjects


def create_stat_columns(columns=None):
    if not columns:
        columns = DEFAULT_COLUMNS
    for C in columns:
        C()


def create_subjects(subjects):
    """Initialize subject columns, but not make them

    arguments:
    subjects -- iterable containing subjects as strings
    """
    for s in subjects:
        DataColumn(s)


if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active

    subjects = ['AP Euro', 'AP Bio', 'Band', 'Dab', 'dab2']

    create_stat_columns()
    create_subjects(subjects)
    Column.make_all(ws)
    print(Column.get_column_strings())

    wb.save('test.xlsx')
