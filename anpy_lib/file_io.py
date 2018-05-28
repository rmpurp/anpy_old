#!/usr/bin/env python3

import json
from openpyxl import Workbook
from openpyxl import load_workbook as load
from anpy_lib.utils import get_most_recent_monday

TEMP_SHEET_NAME = 'temp_sheet_anpy'
JSON_FILE_NAME = 'data.json'

WORKSHEET_CREATED = True


def load_workbook(path):
    try:
        wb = load(path)
    except FileNotFoundError:
        wb = Workbook()
        wb.active.title = TEMP_SHEET_NAME
    return wb


def get_relevant_worksheet(workbook, date=None):
    reference_date = str(get_most_recent_monday(date))
    status = WORKSHEET_CREATED
    if reference_date not in workbook.sheetnames:
        workbook.create_sheet(title=reference_date)
        status = not WORKSHEET_CREATED

    if TEMP_SHEET_NAME in workbook.sheetnames:
        del workbook[TEMP_SHEET_NAME]
    return status, workbook[reference_date]


def save_file(wb, path):
    wb.save(path)


def load_json(path):
    with open(path) as f:
        return json.load(f)
