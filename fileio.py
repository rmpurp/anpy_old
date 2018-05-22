import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import itertools as it
from openpyxl.worksheet.cell_range import CellRange

NUM_BODY_ITEMS = 7

class Column:
    def __init__(self, column, title):
        self.title = title
        self.column = column

    def make(self, ws):
        ws.cell(row=1, column=self.column, value=self.title)
        for row, item in enumerate(self._get_body(NUM_BODY_ITEMS), 2):
            cell = ws.cell(row=row, column=self.column, value=item)
            self._body_cell_op(cell)
        ws.cell(row=row + 1, column=self.column, value=self._get_footer())

    def _get_body_item(self, item_num):
        return item_num

    def _body_cell_op(self, cell):
        pass

    def _get_body(self, num_items):
        i = 1
        while i <= num_items:
            yield self._get_body_item(i)
            i += 1

    def _get_footer(self):
        range = CellRange(min_col=self.column,
                          max_col=self.column,
                          min_row=2,
                          max_row=1 + NUM_BODY_ITEMS)
        return '=AVERAGE({})'.format(str(range))

class DateColumn(Column):
    def __init__(self, column, start_date=None):
        if not start_date:
            start_date = get_most_recent_monday()
        self.start_date = start_date
        super().__init__(column, 'Date')

    def _get_body_item(self, item_num):
        if item_num == 1:
            return self.start_date
        else:
            return '= {}{} + 1'.format(get_column_letter(self.column), item_num)

    def _body_cell_op(self, cell):
        cell.number_format = 'YYYY-MM-DD'

    def _get_footer(self):
        return 'Averages:'

class DefaultValueColumn(Column):
    def __init__(self, column, title, default_value):
        self.default_value = default_value
        super().__init__(column, title)

    def _get_body_item(self, item_num):
        return self.default_value

class TimeColumn(DefaultValueColumn):
    def _get_footer(self):
        return 'N/A'

class TimeTotalColumn(Column):
    def __init__(self, column, time_started_col, time_ended_col):
        self.time_started_col = get_column_letter(time_started_col)
        self.time_ended_col = get_column_letter(time_ended_col)
        super().__init__(column, 'Time Total (H)')

    def _get_body_item(self, item_num):
        row = item_num + 1
        template = '=IF({0}{1}="N/A","N/A",' \
                + ' (MOD(24 + (60 * HOUR({2}{1}) - 60 * HOUR({0}{1})' \
                + ' + MINUTE({2}{1}) - MINUTE({0}{1})) / 60, 24)))'
        return template.format(self.time_started_col, row, self.time_ended_col)

class TimeWorkingColumn(Column):
    def __init__(self, column, subject_start_col, num_subjects):
        self.start_col_idx = subject_start_col
        self.end_col_idx = subject_start_col + num_subjects - 1
        super().__init__(column, 'Time Working (H)')

    def _get_body_item(self, item_num):
        row = item_num + 1
        template = '=IF(SUM({0})=0,"N/A",SUM({0})/60)'
        cell_range = CellRange(min_col=self.start_col_idx,
                          max_col=self.end_col_idx,
                          min_row=row,
                          max_row=row)
        return template.format(cell_range)

class EfficiencyColumn(Column):
    def __init__(self, column, time_total_col, time_working_col):
        self.time_total_col = time_total_col
        self.time_working_col = time_working_col
        super().__init__(column, '% Efficiency') 

    def _get_body_item(self, item_num):
        row = item_num + 1
        template = '=IF({2}{1}="N/A","N/A",IFERROR({2}{1}/({0}{1}*0.75),0))'
        return template.format(get_column_letter(self.time_total_col), 
                               row,
                               get_column_letter(self.time_working_col))

    def _get_footer(self):
        total_range = CellRange(min_col=self.time_total_col,
                                max_col=self.time_total_col,
                                min_row=2,
                                max_row=NUM_BODY_ITEMS + 1)
        my_range = CellRange(min_col=self.column,
                             max_col=self.column,
                             min_row=2,
                             max_row=NUM_BODY_ITEMS + 1)
        template = '=SUMPRODUCT({0},{1}) / SUM({0})'
        return template.format(total_range, my_range)

subjects = ['AP Euro', 'AP Bio', 'Band', 'Dab', 'dab2']

def get_subjects(ws, num_titles):
    subjects = []
    index = num_titles + 1
    cell = ws.cell(row=1, column=index)
    while cell.value:
        subjects.append(cell.value)
        index += 1
        cell = ws.cell(row=1, column=index)
    return subjects

def get_most_recent_monday(date=None):
    if not date:
        date = datetime.date.today()
    return date - datetime.timedelta(days=date.weekday())
        
def create_subject_columns(start_column, subject_names, worksheet):
    for column, name in enumerate(subject_names, start_column):
        c = DefaultValueColumn(column, name, 0)
        c.make(worksheet)

def create_stat_columns(ws):
    DateColumn(1).make(ws)
    TimeColumn(2, 'Time Started', 'N/A').make(ws)
    TimeColumn(3, 'Time Ended', 'N/A').make(ws)
    TimeTotalColumn(4, 2, 3).make(ws)
    TimeWorkingColumn(5, subject_start_col=7, num_subjects=5).make(ws)
    EfficiencyColumn(6, time_total_col=4, time_working_col=5).make(ws)

def make_worksheet(ws):
    create_stat_columns(ws)
    create_subject_columns(7, subjects, ws)

wb = Workbook()
ws = wb.active
make_worksheet(ws)

wb.save('test.xlsx')
