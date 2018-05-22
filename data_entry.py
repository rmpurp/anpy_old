import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import column_creation

class Datum:
    def __init__(self, date, value, column_class):
        self.date = date
        self.value = value
        self.column_class = column_class

    def add(self, ws, columns):
        col = self.get_col(columns)
        row = self.get_row(ws, columns)
        ws.cell(row=row, column=col, value=self.value)

    def get_row(self, ws, columns):
        start_row = 2
        date_column = get_col_from_class(columns, column_creation.DateColumn)
        base_date = ws.cell(row=start_row, column=date_column).value
        date_difference = (self.date - base_date).days
        return start_row + date_difference

    def get_col(self, columns):
        return get_col_from_class(columns, self.column_class)

def get_col_from_class(columns, col_class):
    return columns.index(col_class.__name__) + 1

if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    
    subjects = ['AP Euro', 'AP Bio', 'Band', 'Dab', 'dab2']
    
    column_creation.create_stat_columns()
    column_creation.create_subjects(subjects)
    column_creation.Column.make_all(ws)
    columns = column_creation.Column.get_column_strings()
    time_started = Datum(datetime.date(2018, 5, 22), datetime.time(13, 1),
            column_creation.TimeStartedColumn)
    time_ended = Datum(datetime.date(2018, 5, 22), datetime.time(15, 1),
            column_creation.TimeEndedColumn)
 
    time_started.add(ws, columns)
    time_ended.add(ws, columns)
    
    wb.save('test.xlsx')
